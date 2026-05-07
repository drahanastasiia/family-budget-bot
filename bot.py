import datetime as dt
import logging
import os
from datetime import datetime
from typing import Optional

import pytz
from dotenv import load_dotenv
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    PicklePersistence,
    filters,
)

import database

load_dotenv()

logging.basicConfig(format='%(asctime)s  %(levelname)s  %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# ── constants ────────────────────────────────────────────────────────────────

CATEGORIES: dict[str, str] = {
    'food':          '🍔 Їжа',
    'transport':     '🚗 Транспорт',
    'utilities':     '🔄 Регулярні платежі',
    'entertainment': '🎮 Розваги',
    'health':        '💊 Здоров\'я',
    'clothing':      '👕 Одяг',
    'beauty':        '💄 Краса',
    'other':         '📦 Інше',
}

SUBCATEGORIES: dict[str, list[tuple[str, str]]] = {
    'food': [
        ('groceries',  '🛒 Продукти'),
        ('delivery',   '🚚 Доставка'),
        ('cafe',       '☕ Кафе / ресторан'),
        ('cigarettes', '🚬 Цигарки'),
    ],
    'transport': [
        ('taxi',    '🚕 Таксі'),
        ('public',  '🚌 Громадський транспорт'),
        ('fuel',    '⛽ Пальне'),
        ('parking', '🅿️ Парковка'),
    ],
    'utilities': [
        ('communal',      '💡 Комуналка'),
        ('internet',      '🌐 Інтернет'),
        ('mobile',        '📱 Мобільний зв\'язок'),
        ('water',         '💧 Бутильована вода'),
        ('subscriptions', '📺 Підписки'),
    ],
    'entertainment': [
        ('bar',    '🍺 Бар / клуб'),
        ('cinema', '🎬 Кіно / театр'),
        ('games',  '🎮 Ігри / апки'),
    ],
    'health': [
        ('medicine', '💊 Ліки'),
        ('doctor',   '🏥 Лікар / аналізи'),
        ('sport',    '🏋️ Спорт / зал'),
    ],
    'clothing': [
        ('clothes',     '👕 Одяг'),
        ('shoes',       '👟 Взуття'),
        ('accessories', '👜 Аксесуари'),
    ],
    'beauty': [
        ('manicure', '💅 Манікюр'),
        ('pedicure', '🦶 Педікюр'),
        ('haircut',  '✂️ Стрижка'),
        ('other',    '🪞 Інше'),
    ],
    'other': [
        ('gifts',    '🎁 Подарунки'),
        ('transfer', '💸 Переказ / подяка'),
    ],
}

MONTH_UA = [
    '', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
    'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень',
]

MONTH_UA_GEN = [
    '', 'Січня', 'Лютого', 'Березня', 'Квітня', 'Травня', 'Червня',
    'Липня', 'Серпня', 'Вересня', 'Жовтня', 'Листопада', 'Грудня',
]

S_IDLE        = None
S_WAIT_AMOUNT = 'wait_amount'
S_WAIT_SUBCAT = 'wait_subcat'
S_WAIT_DESC   = 'wait_desc'

KYIV_TZ = pytz.timezone('Europe/Kyiv')

EVENING_MESSAGES = [
    'Хто знає куди йдуть гроші сьогодні — той вирішує, куди вони підуть завтра 🏡',
    'Кожен записаний день — це ще один крок до вашої великої покупки 🎯',
    'Ви будуєте фінансову картину — і це вже більше, ніж роблять більшість 💪',
    'Маленькі звички сьогодні = великі можливості завтра ✨',
    'Ваша мрія стає реальнішою з кожним записом 🚀',
]

# ── keyboards ────────────────────────────────────────────────────────────────

def _main_menu() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton('➕ Додати витрату',   callback_data='menu_add'),
            InlineKeyboardButton('📊 Звіт',             callback_data='menu_report'),
        ],
        [
            InlineKeyboardButton('📋 Список витрат',    callback_data='menu_list'),
            InlineKeyboardButton('↩️ Скасувати останню', callback_data='menu_undo'),
        ],
    ])


def _category_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        InlineKeyboardButton(label, callback_data=f'cat_{key}')
        for key, label in CATEGORIES.items()
    ]
    rows = [buttons[i:i + 2] for i in range(0, len(buttons), 2)]
    return InlineKeyboardMarkup(rows)


def _subcategory_keyboard(category_key: str) -> InlineKeyboardMarkup:
    subs = SUBCATEGORIES.get(category_key, [])
    buttons = [InlineKeyboardButton(label, callback_data=f'subcat_{key}') for key, label in subs]
    rows = [buttons[i:i + 2] for i in range(0, len(buttons), 2)]
    rows.append([InlineKeyboardButton('⏭ Пропустити', callback_data='skip_subcat')])
    return InlineKeyboardMarkup(rows)


def _subcat_label(category: str, subcategory: str) -> str:
    if not subcategory:
        return ''
    return dict(SUBCATEGORIES.get(category, [])).get(subcategory, subcategory)


def _skip_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton('⏭ Пропустити', callback_data='skip_desc'),
        InlineKeyboardButton('❌ Скасувати',  callback_data='cancel'),
    ]])


def _report_nav(month: int, year: int) -> InlineKeyboardMarkup:
    now = datetime.now(KYIV_TZ)
    prev_month = month - 1 or 12
    prev_year  = year if month > 1 else year - 1
    next_month = month % 12 + 1
    next_year  = year if month < 12 else year + 1
    is_future  = (next_year > now.year) or (next_year == now.year and next_month > now.month)

    row = [InlineKeyboardButton(
        f'◀️ {MONTH_UA[prev_month]}', callback_data=f'report_{prev_month}_{prev_year}'
    )]
    if not is_future:
        row.append(InlineKeyboardButton(
            f'{MONTH_UA[next_month]} ▶️', callback_data=f'report_{next_month}_{next_year}'
        ))
    return InlineKeyboardMarkup([row, [InlineKeyboardButton('🏠 Меню', callback_data='menu_home')]])

# ── report builder ───────────────────────────────────────────────────────────

def _prev_month(month: int, year: int):
    return (month - 1 or 12), (year if month > 1 else year - 1)


def _build_report(month: int, year: int) -> Optional[str]:
    expenses = database.get_monthly_expenses(month, year)
    if not expenses:
        return None

    total = sum(e['amount'] for e in expenses)

    by_cat: dict[str, float]  = {}
    by_user: dict[str, float] = {}
    for e in expenses:
        by_cat[e['category']]  = by_cat.get(e['category'], 0)  + e['amount']
        by_user[e['username']] = by_user.get(e['username'], 0) + e['amount']

    # comparison with previous month
    pm, py      = _prev_month(month, year)
    prev_exp    = database.get_monthly_expenses(pm, py)
    prev_total  = sum(e['amount'] for e in prev_exp) if prev_exp else None
    prev_by_cat: dict[str, float] = {}
    for e in prev_exp:
        prev_by_cat[e['category']] = prev_by_cat.get(e['category'], 0) + e['amount']

    def diff_str(cur, prev) -> str:
        if not prev:
            return ''
        pct = (cur - prev) / prev * 100
        arrow = '📈' if pct > 0 else '📉'
        return f'  {arrow} {pct:+.0f}% vs {MONTH_UA_GEN[pm]}'

    total_diff = diff_str(total, prev_total)

    by_subcat: dict[str, dict[str, float]] = {}
    # (category, subcategory) → list of expenses with descriptions
    commented: dict[tuple, list] = {}
    for e in expenses:
        cat = e['category']
        sub = e.get('subcategory') or ''
        if sub:
            by_subcat.setdefault(cat, {})
            by_subcat[cat][sub] = by_subcat[cat].get(sub, 0) + e['amount']
        if e.get('description'):
            key_cs = (cat, sub)
            commented.setdefault(key_cs, [])
            commented[key_cs].append(e)

    lines = [
        f'📊 <b>Звіт за {MONTH_UA[month]} {year}</b>',
        '',
        f'💰 <b>Загальна сума: {total:,.2f} грн</b>{total_diff}',
        '',
        '📂 <b>По категоріях:</b>',
    ]
    for key, amt in sorted(by_cat.items(), key=lambda x: -x[1]):
        pct      = amt / total * 100
        label    = CATEGORIES.get(key, key)
        cat_diff = diff_str(amt, prev_by_cat.get(key))
        lines.append(f'  {label}: <b>{amt:,.2f} грн</b>  ({pct:.0f}%){cat_diff}')
        subs = by_subcat.get(key, {})
        if subs:
            for sub_key, sub_amt in sorted(subs.items(), key=lambda x: -x[1]):
                sub_label = _subcat_label(key, sub_key)
                lines.append(f'    · {sub_label}: {sub_amt:,.0f} грн')
                for c in commented.get((key, sub_key), []):
                    lines.append(f'      └ <i>"{c["description"]}"</i> — {c["amount"]:,.0f} грн  {c["username"]}')
        # expenses with no subcategory but with description
        for c in commented.get((key, ''), []):
            lines.append(f'    └ <i>"{c["description"]}"</i> — {c["amount"]:,.0f} грн  {c["username"]}')

    lines += ['', '👥 <b>По учасниках:</b>']
    for uname, amt in sorted(by_user.items(), key=lambda x: -x[1]):
        lines.append(f'  {uname}: {amt:,.2f} грн')

    if prev_total:
        saved = prev_total - total
        if saved > 0:
            lines += ['', f'🎉 <b>Зекономили {saved:,.2f} грн порівняно з {MONTH_UA_GEN[pm]}!</b>']

    return '\n'.join(lines)


def _build_list(month: int, year: int) -> Optional[str]:
    expenses = database.get_monthly_expenses(month, year)
    if not expenses:
        return None
    lines = [f'📋 <b>Витрати за {MONTH_UA[month]} {year}</b> ({len(expenses)} записів)\n']
    for e in expenses[:20]:
        day  = e['created_at'][8:10]
        cat  = CATEGORIES.get(e['category'], e['category'])
        sub  = e.get('subcategory') or ''
        sub_part = f' · {_subcat_label(e["category"], sub)}' if sub else ''
        desc = f' — {e["description"]}' if e['description'] else ''
        lines.append(f'{day}.{month:02d}  {cat}{sub_part}  <b>{e["amount"]:,.0f} грн</b>{desc}  <i>{e["username"]}</i>')
    if len(expenses) > 20:
        lines.append(f'\n…і ще {len(expenses) - 20} записів')
    return '\n'.join(lines)

# ── excel builder ────────────────────────────────────────────────────────────

def _build_excel(month: int, year: int):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    expenses = database.get_monthly_expenses(month, year)
    if not expenses:
        return None

    wb = Workbook()

    # ── Sheet 1: by day ──
    ws1 = wb.active
    ws1.title = 'По днях'
    headers = ['Дата', 'Хто', 'Категорія', 'Підкатегорія', 'Сума, грн', 'Коментар']
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')

    for e in reversed(expenses):
        date = f"{e['created_at'][8:10]}.{e['created_at'][5:7]}.{e['created_at'][:4]}"
        cat  = CATEGORIES.get(e['category'], e['category'])
        sub  = _subcat_label(e['category'], e.get('subcategory') or '')
        ws1.append([date, e['username'], cat, sub, e['amount'], e.get('description') or ''])

    total_row = ['', '', '', 'РАЗОМ:', sum(e['amount'] for e in expenses), '']
    ws1.append(total_row)
    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 30

    # ── Sheet 2: summary by category ──
    ws2 = wb.create_sheet('По категоріях')
    ws2.append(['Категорія', 'Підкатегорія', 'Сума, грн'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')

    by_cat: dict[str, float] = {}
    by_subcat: dict[str, dict[str, float]] = {}
    for e in expenses:
        cat = e['category']
        sub = e.get('subcategory') or ''
        by_cat[cat] = by_cat.get(cat, 0) + e['amount']
        by_subcat.setdefault(cat, {})
        by_subcat[cat][sub] = by_subcat[cat].get(sub, 0) + e['amount']

    for cat_key, cat_amt in sorted(by_cat.items(), key=lambda x: -x[1]):
        cat_label = CATEGORIES.get(cat_key, cat_key)
        row = ws2.max_row + 1
        ws2.append([cat_label, '', cat_amt])
        ws2.cell(row, 1).font = Font(bold=True)
        for sub_key, sub_amt in sorted(by_subcat[cat_key].items(), key=lambda x: -x[1]):
            sub_label = _subcat_label(cat_key, sub_key) if sub_key else '—'
            ws2.append(['', sub_label, sub_amt])

    ws2.append(['РАЗОМ', '', sum(e['amount'] for e in expenses)])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── helpers ──────────────────────────────────────────────────────────────────

def _display_name(user) -> str:
    return f'@{user.username}' if user.username else user.first_name


def _state(context: ContextTypes.DEFAULT_TYPE):
    return context.user_data.get('state', S_IDLE)


def _reset(context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()

# ── command handlers ─────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _reset(context)
    user = update.effective_user
    database.upsert_user(user.id, _display_name(user))
    await update.message.reply_text(
        f'👋 Привіт, <b>{user.first_name}</b>!\n\n'
        '🏡 Цей бот допомагає вам з партнером зрозуміти, куди йдуть гроші — '
        'щоб впевнено планувати великі покупки й досягати спільних цілей.\n\n'
        'Кожен вносить витрати зі свого телефону, а бот збирає загальну картину.\n\n'
        'Що робимо?',
        parse_mode='HTML',
        reply_markup=_main_menu(),
    )


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        '<b>📖 Команди:</b>\n'
        '/start — головне меню\n'
        '/add — додати витрату\n'
        '/report — звіт за поточний місяць\n'
        '/list — список витрат\n'
        '/cancel — скасувати поточну дію\n\n'
        '💡 Просто поділись ботом з рідними — кожен може додавати витрати.',
        parse_mode='HTML',
    )







async def cmd_fixcats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    results = []
    with database._db() as conn:
        # Парковка → transport/parking
        row = conn.execute(
            "SELECT id FROM expenses WHERE username='@Lex8228' AND description='Парковка'"
        ).fetchone()
        if row:
            database.update_expense_subcategory(row['id'], 'parking')
            results.append('✅ Парковка → 🅿️ Парковка')

        # Спасибо → other/transfer
        row = conn.execute(
            "SELECT id FROM expenses WHERE username='@Lex8228' AND description='«Спасибо»'"
        ).fetchone()
        if row:
            database.update_expense_category(row['id'], 'other', 'transfer')
            results.append('✅ «Спасибо» → 💸 Переказ / подяка')

        # Педікюр → beauty/pedicure
        row = conn.execute(
            "SELECT id FROM expenses WHERE username='@ana_drahan' AND description LIKE '%педікюр%'"
        ).fetchone()
        if row:
            database.update_expense_category(row['id'], 'beauty', 'pedicure')
            results.append('✅ Педікюр → 💄 Краса / 🦶 Педікюр')

    await update.message.reply_text('\n'.join(results) if results else '❌ Записи не знайдено.')


async def cmd_dumpdesc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    with database._db() as conn:
        rows = conn.execute(
            "SELECT category, subcategory, description FROM expenses "
            "WHERE description != '' AND description != 'Відновлено' "
            "ORDER BY category, subcategory"
        ).fetchall()
    if not rows:
        await update.message.reply_text('Описів немає.')
        return
    lines = []
    for r in rows:
        cat = CATEGORIES.get(r['category'], r['category'])
        sub = _subcat_label(r['category'], r['subcategory']) if r['subcategory'] else '—'
        lines.append(f'{cat} / {sub}: {r["description"]}')
    await update.message.reply_text('\n'.join(lines))


async def cmd_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _reset(context)
    await update.message.reply_text('❌ Скасовано.', reply_markup=_main_menu())


async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    now  = datetime.now(KYIV_TZ)
    await _send_report(update.message.reply_text, now.month, now.year)


async def cmd_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    now = datetime.now(KYIV_TZ)
    buf = _build_excel(now.month, now.year)
    if not buf:
        await update.message.reply_text('📋 Витрат цього місяця ще немає.')
        return
    filename = f'звіт_{MONTH_UA[now.month].lower()}_{now.year}.xlsx'
    await update.message.reply_document(document=buf, filename=filename)


async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    now  = datetime.now(KYIV_TZ)
    text = _build_list(now.month, now.year)
    await update.message.reply_text(
        text if text else '📋 Витрат цього місяця ще немає.',
        parse_mode='HTML' if text else None,
        reply_markup=_main_menu(),
    )

# ── add flow ──────────────────────────────────────────────────────────────────

async def _ask_amount(send_fn, context: ContextTypes.DEFAULT_TYPE):
    _reset(context)
    context.user_data['state'] = S_WAIT_AMOUNT
    await send_fn('💵 Введи суму витрати (наприклад: 250 або 1500.50):')


async def _save_and_confirm(chat_id: int, user, context: ContextTypes.DEFAULT_TYPE, description: str):
    amount     = context.user_data['amount']
    category   = context.user_data['category']
    subcategory = context.user_data.get('subcategory', '')
    _reset(context)
    database.add_expense(user.id, _display_name(user), amount, category, subcategory, description)

    sub_line = f'\n🏷 {_subcat_label(category, subcategory)}' if subcategory else ''
    text = (
        '✅ <b>Витрату збережено!</b>\n\n'
        f'💵 Сума: <b>{amount:,.2f} грн</b>\n'
        f'📂 Категорія: {CATEGORIES[category]}{sub_line}\n'
        + (f'📝 Опис: {description}' if description else '')
    )
    await context.bot.send_message(
        chat_id=chat_id,
        text=text,
        parse_mode='HTML',
        reply_markup=_main_menu(),
    )

# ── report helpers ────────────────────────────────────────────────────────────

async def _send_report(send_fn, month: int, year: int):
    text = _build_report(month, year)
    await send_fn(
        text if text else f'📊 Витрат за {MONTH_UA[month]} {year} ще немає.',
        parse_mode='HTML' if text else None,
        reply_markup=_report_nav(month, year),
    )

# ── text message handler ──────────────────────────────────────────────────────

async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text  = update.message.text.strip()
    state = _state(context)

    if state == S_WAIT_AMOUNT:
        raw = text.replace(',', '.')
        try:
            amount = float(raw)
            if amount <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text(
                '❌ Не розумію. Введи число, наприклад: 250 або 1500.50'
            )
            return
        context.user_data['amount'] = amount
        context.user_data['state']  = None
        await update.message.reply_text('📂 Вибери категорію:', reply_markup=_category_keyboard())

    elif state == S_WAIT_DESC:
        await _save_and_confirm(update.effective_chat.id, update.effective_user, context, text)

    else:
        await update.message.reply_text(
            'Скористайся меню або введи /add щоб додати витрату.',
            reply_markup=_main_menu(),
        )

# ── callback query handler ────────────────────────────────────────────────────

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data  = query.data
    await query.answer()

    # ── navigation ──
    if data == 'menu_home':
        await query.message.reply_text('Головне меню:', reply_markup=_main_menu())

    elif data == 'menu_add':
        await _ask_amount(query.message.reply_text, context)

    elif data == 'menu_report':
        now = datetime.now(KYIV_TZ)
        await _send_report(query.message.reply_text, now.month, now.year)

    elif data.startswith('report_'):
        _, m, y = data.split('_')
        await _send_report(query.message.reply_text, int(m), int(y))

    elif data == 'menu_list':
        now  = datetime.now(KYIV_TZ)
        text = _build_list(now.month, now.year)
        await query.message.reply_text(
            text if text else '📋 Витрат цього місяця ще немає.',
            parse_mode='HTML' if text else None,
            reply_markup=_main_menu(),
        )

    # ── undo last expense ──
    elif data == 'menu_undo':
        expenses = database.get_monthly_expenses(
            datetime.now(KYIV_TZ).month, datetime.now(KYIV_TZ).year
        )
        user_expenses = [e for e in expenses if e['user_id'] == update.effective_user.id]
        if not user_expenses:
            await query.message.reply_text(
                '📋 У тебе ще немає витрат цього місяця.', reply_markup=_main_menu()
            )
            return
        last = user_expenses[0]
        cat  = CATEGORIES.get(last['category'], last['category'])
        desc = f'\n📝 {last["description"]}' if last['description'] else ''
        await query.message.reply_text(
            f'↩️ <b>Скасувати останню витрату?</b>\n\n'
            f'💵 {last["amount"]:,.2f} грн  {cat}{desc}\n'
            f'🕐 {last["created_at"][8:10]}.{last["created_at"][5:7]}',
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton('🗑 Так, видалити', callback_data=f'del_yes_{last["id"]}'),
                InlineKeyboardButton('❌ Ні', callback_data='menu_home'),
            ]]),
        )

    elif data.startswith('del_yes_'):
        expense_id = int(data.split('_')[2])
        expense    = database.get_expense_by_id(expense_id)
        if expense and expense['user_id'] == update.effective_user.id:
            database.delete_expense(expense_id)
            await query.message.reply_text(
                '✅ Витрату видалено.', reply_markup=_main_menu()
            )
        else:
            await query.message.reply_text(
                '⚠️ Не вдалося видалити.', reply_markup=_main_menu()
            )

    elif data == 'cancel':
        _reset(context)
        await query.message.reply_text('❌ Скасовано.', reply_markup=_main_menu())

    # ── category selection ──
    elif data.startswith('cat_'):
        if context.user_data.get('amount') is None:
            await query.message.reply_text(
                '⚠️ Спочатку введи суму. Натисни /add.', reply_markup=_main_menu()
            )
            return
        key = data[4:]
        context.user_data['category'] = key
        context.user_data['state']    = S_WAIT_SUBCAT
        await query.message.reply_text(
            f'Категорія: {CATEGORIES[key]}\n\n🏷 Уточни підкатегорію (або пропусти):',
            reply_markup=_subcategory_keyboard(key),
        )

    elif data.startswith('subcat_'):
        sub_key = data[7:]
        context.user_data['subcategory'] = sub_key
        context.user_data['state']       = S_WAIT_DESC
        await query.message.reply_text(
            '📝 Додай опис (або пропусти):',
            reply_markup=_skip_keyboard(),
        )

    elif data == 'skip_subcat':
        context.user_data['subcategory'] = ''
        context.user_data['state']       = S_WAIT_DESC
        await query.message.reply_text(
            '📝 Додай опис (або пропусти):',
            reply_markup=_skip_keyboard(),
        )

    elif data == 'skip_desc':
        if context.user_data.get('category') is None:
            await query.message.reply_text('⚠️ Щось пішло не так. Спробуй /add ще раз.')
            _reset(context)
            return
        await _save_and_confirm(update.effective_chat.id, update.effective_user, context, '')

# ── scheduled jobs ────────────────────────────────────────────────────────────

async def evening_reminder(context: ContextTypes.DEFAULT_TYPE):
    import random
    if database.get_today_total() > 0:
        return  # витрати сьогодні вносились — не турбуємо

    text = (
        '🌙 <b>Добрий вечір!</b>\n\n'
        'Сьогодні ще не було записів — можливо щось випустили? 😊\n\n'
        f'{random.choice(EVENING_MESSAGES)}'
    )
    for user_id in database.get_all_user_ids():
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=text,
                parse_mode='HTML',
                reply_markup=_main_menu(),
            )
        except Exception as exc:
            logger.warning('Cannot send reminder to %s: %s', user_id, exc)


async def auto_monthly_report(context: ContextTypes.DEFAULT_TYPE):
    now   = datetime.now(KYIV_TZ)
    month = now.month - 1 or 12
    year  = now.year if now.month > 1 else now.year - 1
    text  = _build_report(month, year)
    if not text:
        return
    full     = f'🔔 <b>Автоматичний місячний звіт</b>\n\n{text}'
    buf      = _build_excel(month, year)
    filename = f'звіт_{MONTH_UA[month].lower()}_{year}.xlsx'
    for user_id in database.get_all_user_ids():
        try:
            await context.bot.send_message(chat_id=user_id, text=full, parse_mode='HTML')
            if buf:
                buf.seek(0)
                await context.bot.send_document(chat_id=user_id, document=buf, filename=filename)
        except Exception as exc:
            logger.warning('Cannot send auto-report to %s: %s', user_id, exc)

# ── error handler ─────────────────────────────────────────────────────────────

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error('Unhandled exception:', exc_info=context.error)

# ── main ─────────────────────────────────────────────────────────────────────

def main():
    database.init_db()

    token = os.getenv('BOT_TOKEN')
    if not token:
        raise RuntimeError('BOT_TOKEN is not set.')

    data_dir    = os.getenv('DATA_DIR', '.')
    persistence = PicklePersistence(filepath=os.path.join(data_dir, 'conversations.pickle'))
    app         = Application.builder().token(token).persistence(persistence).build()

    app.add_handler(CommandHandler('start',    cmd_start))
    app.add_handler(CommandHandler('help',     cmd_help))
    app.add_handler(CommandHandler('cancel',   cmd_cancel))
    app.add_handler(CommandHandler('dumpdesc', cmd_dumpdesc))
    app.add_handler(CommandHandler('fixcats',  cmd_fixcats))
    app.add_handler(CommandHandler('report', cmd_report))
    app.add_handler(CommandHandler('list',   cmd_list))
    app.add_handler(CommandHandler('excel',  cmd_excel))

    app.add_handler(CommandHandler('add',    lambda u, c: _ask_amount(u.message.reply_text, c)))
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    app.add_error_handler(error_handler)

    app.job_queue.run_daily(
        evening_reminder,
        time=dt.time(21, 0, tzinfo=KYIV_TZ),
    )
    app.job_queue.run_monthly(
        auto_monthly_report,
        when=dt.time(9, 0, tzinfo=KYIV_TZ),
        day=1,
    )

    async def post_init(application: Application):
        await application.bot.set_my_commands([
            ('start',  'Головне меню'),
            ('add',    'Додати витрату'),
            ('report', 'Звіт за поточний місяць'),
            ('list',   'Список витрат'),
            ('cancel', 'Скасувати поточну дію'),
            ('help',   'Допомога'),
        ])

    app.post_init = post_init

    logger.info('Bot started.')
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == '__main__':
    main()
